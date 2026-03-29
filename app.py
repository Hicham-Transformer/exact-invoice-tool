from __future__ import annotations

import io
import os
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import List, Optional, Dict, Any

import pandas as pd
import requests
from flask import (
    Flask,
    Response,
    redirect,
    render_template_string,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "exact-pdf-mission-freight-tool")

CLIENT_ID = os.environ.get("EXACT_CLIENT_ID")
CLIENT_SECRET = os.environ.get("EXACT_CLIENT_SECRET")
REDIRECT_URI = os.environ.get("EXACT_REDIRECT_URI")

AUTH_URL = "https://start.exactonline.nl/api/oauth2/auth"
TOKEN_URL = "https://start.exactonline.nl/api/oauth2/token"
BASE_URL = "https://start.exactonline.nl/api/v1"

TARGET_SUPPLIER = "mission freight"

CHARGE_KEYWORDS = [
    "import warehouse charges",
    "handling",
    "handling charges",
    "handling fee",
]


@dataclass
class PdfInvoiceResult:
    bestandsnaam: str
    factuurnummer: Optional[str]
    awb_nummer: Optional[str]
    totaal_kg: Optional[float]
    charges_eur: Optional[float]
    prijs_per_kg_eur: Optional[float]
    status: str


def normalize_supplier(text: str) -> str:
    return (text or "").strip().lower()


def safe_json(response) -> Optional[dict]:
    text = response.text or ""
    if not text.strip():
        return None
    try:
        return response.json()
    except Exception:
        return None


def exact_date_to_text(value: Any) -> str:
    if not value:
        return ""
    text = str(value)
    m = re.search(r"/Date\((\d+)", text)
    if m:
        try:
            return pd.to_datetime(int(m.group(1)), unit="ms").strftime("%Y-%m-%d")
        except Exception:
            return text
    return text


def get_current_division(headers: dict) -> str:
    res = requests.get(f"{BASE_URL}/current/Me", headers=headers, timeout=30)

    if res.status_code != 200:
        raise RuntimeError(f"Fout bij ophalen division: {res.text}")

    data = safe_json(res)
    if data:
        try:
            return str(data["d"]["results"][0]["CurrentDivision"])
        except Exception:
            pass

    text = res.text or ""
    match = re.search(r"<d:CurrentDivision>(\d+)</d:CurrentDivision>", text)
    if match:
        return match.group(1)

    raise RuntimeError(f"Division niet gevonden: {text[:300]}")


def get_all_purchase_entries(headers: dict, division: str) -> List[Dict[str, Any]]:
    url = f"{BASE_URL}/{division}/purchaseentry/PurchaseEntries?$top=100"
    all_results: List[Dict[str, Any]] = []

    while url:
        res = requests.get(url, headers=headers, timeout=60)

        if res.status_code != 200:
            raise RuntimeError(f"Exact fout: {res.text}")

        data = safe_json(res)
        if not data:
            raise RuntimeError(f"Geen JSON van Exact ontvangen: {res.text[:300]}")

        d = data.get("d", {})
        results = d.get("results", []) if isinstance(d, dict) else []
        all_results.extend(results)

        url = d.get("__next") if isinstance(d, dict) else None

    return all_results


def normalize_spaces(text: str) -> str:
    text = text.replace("\xa0", " ").replace("\u200b", " ")
    text = re.sub(r"[ \t]+", " ", text)
    return text


def parse_decimal_eu(value: str) -> Optional[Decimal]:
    cleaned = value.strip().replace("EUR", "").replace("€", "").replace(" ", "")
    if not cleaned:
        return None

    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")

    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def extract_text_from_pdf_bytes(data: bytes) -> str:
    if fitz is None:
        raise RuntimeError("PyMuPDF is niet beschikbaar. Voeg pymupdf toe aan requirements.txt.")
    with fitz.open(stream=data, filetype="pdf") as doc:
        return "\n".join(page.get_text("text") for page in doc)


def extract_words_from_pdf_bytes(data: bytes):
    if fitz is None:
        raise RuntimeError("PyMuPDF is niet beschikbaar. Voeg pymupdf toe aan requirements.txt.")
    words = []
    with fitz.open(stream=data, filetype="pdf") as doc:
        for page_index, page in enumerate(doc):
            for w in page.get_text("words"):
                words.append((page_index, *w))
    return words


def find_invoice_number(text: str) -> Optional[str]:
    patterns = [
        r"FACTUURNUMMER\s+([0-9]{5,})",
        r"FACTUURNUMMER\s*[:\-]?\s*([A-Z0-9\-]{5,})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return None


def find_awb_number(text: str) -> Optional[str]:
    patterns = [
        r"AWB\s*NUMMER\s*[:\-]?\s*([0-9]{3}-[0-9]{8,})",
        r"\b([0-9]{3}-[0-9]{8,})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return None


def find_total_weight_kg_from_words(words) -> Optional[Decimal]:
    if not words:
        return None

    bruto_words = []
    for item in words:
        page_index, x0, y0, x1, y1, text, block_no, line_no, word_no = item
        if str(text).strip().lower() == "bruto":
            bruto_words.append(item)

    if not bruto_words:
        return None

    bruto_word = sorted(bruto_words, key=lambda w: (w[0], w[3], w[2]))[-1]
    b_page, bx0, by0, bx1, by1, _, _, _, _ = bruto_word
    bruto_center_x = (bx0 + bx1) / 2

    candidates = []
    for item in words:
        page_index, x0, y0, x1, y1, text, block_no, line_no, word_no = item

        if page_index != b_page:
            continue

        token = str(text).strip()
        if not re.fullmatch(r"\d+(?:[.,]\d+)?", token):
            continue

        center_x = (x0 + x1) / 2

        if y0 <= by1:
            continue

        if abs(center_x - bruto_center_x) > 100:
            continue

        value = parse_decimal_eu(token)
        if value is None:
            continue

        candidates.append((y0, value))

    if candidates:
        candidates.sort(key=lambda item: item[0])
        return candidates[0][1]

    return None


def sum_relevant_charges(text: str) -> Optional[Decimal]:
    total = Decimal("0")
    found = False

    lines = [normalize_spaces(line).strip().lower() for line in text.splitlines() if line.strip()]

    for line in lines:
        if not any(keyword in line for keyword in CHARGE_KEYWORDS):
            continue

        amounts = re.findall(r"([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{2}))", line)
        if not amounts:
            continue

        amount = parse_decimal_eu(amounts[-1])
        if amount is not None:
            total += amount
            found = True

    return total if found else None


def parse_pdf_invoice(file_name: str, data: bytes) -> PdfInvoiceResult:
    try:
        text = extract_text_from_pdf_bytes(data)
        words = extract_words_from_pdf_bytes(data)

        factuurnummer = find_invoice_number(text)
        awb_nummer = find_awb_number(text)
        totaal_kg = find_total_weight_kg_from_words(words)
        charges = sum_relevant_charges(text)

        prijs_per_kg = None
        if charges is not None and totaal_kg not in (None, Decimal("0")):
            prijs_per_kg = charges / totaal_kg

        missing = []
        if not factuurnummer:
            missing.append("factuurnummer")
        if not awb_nummer:
            missing.append("AWB nummer")
        if totaal_kg is None:
            missing.append("totaal kg")
        if charges is None:
            missing.append("Import warehouse charges / Handling")

        status = "OK" if not missing else f"Ontbreekt: {', '.join(missing)}"

        return PdfInvoiceResult(
            bestandsnaam=file_name,
            factuurnummer=factuurnummer,
            awb_nummer=awb_nummer,
            totaal_kg=float(totaal_kg) if totaal_kg is not None else None,
            charges_eur=float(charges) if charges is not None else None,
            prijs_per_kg_eur=float(prijs_per_kg) if prijs_per_kg is not None else None,
            status=status,
        )
    except Exception as exc:
        return PdfInvoiceResult(
            bestandsnaam=file_name,
            factuurnummer=None,
            awb_nummer=None,
            totaal_kg=None,
            charges_eur=None,
            prijs_per_kg_eur=None,
            status=f"Fout bij verwerken: {exc}",
        )


def fetch_exact_mission_freight_rows(token: str) -> List[Dict[str, Any]]:
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
    }

    division = get_current_division(headers)
    entries = get_all_purchase_entries(headers, division)

    filtered = [
        e for e in entries
        if TARGET_SUPPLIER in normalize_supplier(e.get("SupplierName") or "")
    ]

    results: List[Dict[str, Any]] = []

    for item in filtered:
        results.append(
            {
                "Bron": "PurchaseEntries",
                "Factuurnummer": str(item.get("InvoiceNumber") or ""),
                "Factuurdatum": exact_date_to_text(item.get("EntryDate", "")),
                "Leverancier": item.get("SupplierName", ""),
                "Exact omschrijving": item.get("Description", ""),
                "Exact totaal DC": item.get("AmountDC", 0),
                "Valuta": item.get("Currency", ""),
                "Exact document id": item.get("EntryID", ""),
                "Exact boekingsnummer": item.get("EntryNumber", ""),
                "Exact status": item.get("Status", ""),
            }
        )

    return results


def merge_exact_and_pdf(exact_rows: List[Dict[str, Any]], pdf_results: List[PdfInvoiceResult]) -> pd.DataFrame:
    pdf_map: Dict[str, PdfInvoiceResult] = {}
    for p in pdf_results:
        if p.factuurnummer:
            pdf_map[str(p.factuurnummer)] = p

    merged_rows: List[Dict[str, Any]] = []
    used_pdf_numbers = set()

    for row in exact_rows:
        factuurnummer = str(row.get("Factuurnummer") or "")
        pdf = pdf_map.get(factuurnummer)
        if pdf:
            used_pdf_numbers.add(factuurnummer)

        merged_rows.append(
            {
                "Factuurnummer": factuurnummer,
                "AWB nummer": pdf.awb_nummer if pdf else "",
                "Factuurdatum": row.get("Factuurdatum", ""),
                "Leverancier": row.get("Leverancier", ""),
                "Totaal kg": pdf.totaal_kg if pdf else None,
                "Charges (EUR)": pdf.charges_eur if pdf else None,
                "Prijs per kg (EUR)": pdf.prijs_per_kg_eur if pdf else None,
                "PDF bestandsnaam": pdf.bestandsnaam if pdf else "",
                "PDF status": pdf.status if pdf else "Geen PDF gekoppeld",
                "Bron": row.get("Bron", ""),
                "Exact document id": row.get("Exact document id", ""),
                "Exact boekingsnummer": row.get("Exact boekingsnummer", ""),
                "Exact omschrijving": row.get("Exact omschrijving", ""),
                "Exact totaal DC": row.get("Exact totaal DC", 0),
                "Valuta": row.get("Valuta", ""),
                "Exact status": row.get("Exact status", ""),
            }
        )

    for p in pdf_results:
        if not p.factuurnummer or p.factuurnummer in used_pdf_numbers:
            continue

        merged_rows.append(
            {
                "Factuurnummer": p.factuurnummer,
                "AWB nummer": p.awb_nummer,
                "Factuurdatum": "",
                "Leverancier": "Mission Freight (alleen PDF)",
                "Totaal kg": p.totaal_kg,
                "Charges (EUR)": p.charges_eur,
                "Prijs per kg (EUR)": p.prijs_per_kg_eur,
                "PDF bestandsnaam": p.bestandsnaam,
                "PDF status": p.status,
                "Bron": "PDF only",
                "Exact document id": "",
                "Exact boekingsnummer": "",
                "Exact omschrijving": "",
                "Exact totaal DC": "",
                "Valuta": "",
                "Exact status": "",
            }
        )

    return pd.DataFrame(merged_rows)


def build_excel_bytes(df: pd.DataFrame, pdf_df: pd.DataFrame, exact_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Samengevoegd"

    headers = list(df.columns)
    ws.append(headers)

    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    for _, row in df.iterrows():
        ws.append([row.get(col) for col in headers])

    widths = {
        "A": 16, "B": 18, "C": 14, "D": 22, "E": 12, "F": 14, "G": 16,
        "H": 24, "I": 18, "J": 18, "K": 18, "L": 24, "M": 14, "N": 12, "O": 14
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws2 = wb.create_sheet("PDF raw")
    if not pdf_df.empty:
        ws2.append(list(pdf_df.columns))
        for _, row in pdf_df.iterrows():
            ws2.append([row.get(col) for col in pdf_df.columns])

    ws3 = wb.create_sheet("Exact raw")
    if not exact_df.empty:
        ws3.append(list(exact_df.columns))
        for _, row in exact_df.iterrows():
            ws3.append([row.get(col) for col in exact_df.columns])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def session_get_df(key: str) -> pd.DataFrame:
    raw = session.get(key)
    if not raw:
        return pd.DataFrame()
    return pd.read_json(io.StringIO(raw))


HTML = """
<!doctype html>
<html lang="nl">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Exact + PDF Mission Freight Tool</title>
  <style>
    body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, sans-serif; background: #f5f7fb; margin: 0; padding: 18px; color: #14213d; }
    .wrap { max-width: 1000px; margin: 0 auto; }
    .card { background: white; border: 1px solid #dbe2f0; border-radius: 18px; padding: 20px; margin-bottom: 16px; }
    h1 { margin: 0 0 10px; font-size: 1.7rem; }
    h2 { margin: 0 0 10px; font-size: 1.25rem; }
    p { color: #5b6475; line-height: 1.5; }
    .btn { display: inline-block; background: #2563eb; color: white; padding: 14px 18px; border-radius: 14px; text-decoration: none; font-weight: 700; border: 0; cursor: pointer; margin-top: 10px; }
    .btn-secondary { background: #eef3ff; color: #2563eb; }
    input[type=file] { width: 100%; margin-top: 12px; font-size: 16px; }
    .ok { color: #117a37; font-weight: 700; }
    .warn { color: #b26a00; font-weight: 700; }
    table { width: 100%; border-collapse: collapse; font-size: 14px; }
    th, td { padding: 10px; border-bottom: 1px solid #dbe2f0; text-align: left; vertical-align: top; }
    th { color: #5b6475; }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <h1>Exact + PDF Mission Freight Tool</h1>
      <p>Stap 1: haal Mission Freight facturen uit Exact.</p>
      <p>Stap 2: upload dezelfde Mission Freight PDF-facturen.</p>
      <p>Stap 3: download één Excel met Exact + PDF gecombineerd.</p>

      {% if connected %}
        <p class="ok">Exact is gekoppeld.</p>
        <a class="btn" href="{{ url_for('fetch_exact') }}">1. Haal Mission Freight uit Exact</a>
      {% else %}
        <a class="btn" href="{{ url_for('login') }}">Login met Exact</a>
      {% endif %}
    </div>

    <div class="card">
      <h2>2. Upload Mission Freight PDF's</h2>
      <form method="post" action="{{ url_for('upload_pdfs') }}" enctype="multipart/form-data">
        <input type="file" name="files" accept="application/pdf" multiple required>
        <button class="btn" type="submit">Verwerk PDF's</button>
      </form>

      {% if can_download %}
        <a class="btn btn-secondary" href="{{ url_for('download_excel') }}">3. Download Excel</a>
      {% endif %}
    </div>

    <div class="card">
      <h2>Status</h2>
      <p>Exact regels: <strong>{{ exact_count }}</strong></p>
      <p>PDF regels: <strong>{{ pdf_count }}</strong></p>
      <p>Samengevoegde regels: <strong>{{ merged_count }}</strong></p>
    </div>

    {% if merged_rows %}
    <div class="card">
      <h2>Voorbeeld resultaten</h2>
      <table>
        <thead>
          <tr>
            <th>Factuurnummer</th>
            <th>AWB</th>
            <th>Factuurdatum</th>
            <th>Leverancier</th>
            <th>KG</th>
            <th>Charges</th>
            <th>Prijs/kg</th>
            <th>PDF status</th>
          </tr>
        </thead>
        <tbody>
          {% for row in merged_rows[:20] %}
          <tr>
            <td>{{ row.get('Factuurnummer', '') }}</td>
            <td>{{ row.get('AWB nummer', '') }}</td>
            <td>{{ row.get('Factuurdatum', '') }}</td>
            <td>{{ row.get('Leverancier', '') }}</td>
            <td>{{ row.get('Totaal kg', '') }}</td>
            <td>{{ row.get('Charges (EUR)', '') }}</td>
            <td>{{ row.get('Prijs per kg (EUR)', '') }}</td>
            <td class="{% if row.get('PDF status','') == 'OK' %}ok{% else %}warn{% endif %}">{{ row.get('PDF status', '') }}</td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
    {% endif %}
  </div>
</body>
</html>
"""


@app.route("/")
def index():
    exact_df = session_get_df("exact_json")
    pdf_df = session_get_df("pdf_json")
    merged_df = session_get_df("merged_json")

    rows = merged_df.to_dict(orient="records") if not merged_df.empty else []

    return render_template_string(
        HTML,
        connected=bool(session.get("access_token")),
        exact_count=len(exact_df),
        pdf_count=len(pdf_df),
        merged_count=len(merged_df),
        can_download=not merged_df.empty,
        merged_rows=rows,
    )


@app.route("/login")
def login():
    if not CLIENT_ID or not CLIENT_SECRET or not REDIRECT_URI:
        return (
            "Environment variables ontbreken. Zet EXACT_CLIENT_ID, "
            "EXACT_CLIENT_SECRET en EXACT_REDIRECT_URI in Render.",
            500,
        )

    url = (
        f"{AUTH_URL}"
        f"?client_id={CLIENT_ID}"
        f"&redirect_uri={REDIRECT_URI}"
        f"&response_type=code"
        f"&scope=exactonlineapi%20offline_access"
    )
    return redirect(url)


@app.route("/callback")
def callback():
    error = request.args.get("error")
    code = request.args.get("code")

    if error:
        return f"Exact fout: {error}", 400

    if not code:
        return "Geen code ontvangen van Exact.", 400

    data = {
        "grant_type": "authorization_code",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "code": code,
        "redirect_uri": REDIRECT_URI,
    }

    res = requests.post(TOKEN_URL, data=data, timeout=30)
    token = safe_json(res)

    if not token:
        return f"Token response niet leesbaar: {res.text}", 400

    access_token = token.get("access_token")
    if not access_token:
        return f"Geen access token ontvangen: {token}", 400

    session["access_token"] = access_token
    session["refresh_token"] = token.get("refresh_token")

    return redirect(url_for("index"))


@app.route("/fetch_exact")
def fetch_exact():
    try:
        token = session.get("access_token")
        if not token:
            return redirect(url_for("login"))

        exact_rows = fetch_exact_mission_freight_rows(token)
        exact_df = pd.DataFrame(exact_rows)
        session["exact_json"] = exact_df.to_json(orient="records")

        pdf_df = session_get_df("pdf_json")
        pdf_results = []
        if not pdf_df.empty:
            for _, r in pdf_df.iterrows():
                pdf_results.append(
                    PdfInvoiceResult(
                        bestandsnaam=r.get("Bestandsnaam"),
                        factuurnummer=r.get("Factuurnummer"),
                        awb_nummer=r.get("AWB nummer"),
                        totaal_kg=r.get("Totaal kg"),
                        charges_eur=r.get("Charges (EUR)"),
                        prijs_per_kg_eur=r.get("Prijs per kg (EUR)"),
                        status=r.get("Status"),
                    )
                )

        merged_df = merge_exact_and_pdf(exact_rows, pdf_results)
        session["merged_json"] = merged_df.to_json(orient="records")

        return redirect(url_for("index"))
    except Exception as e:
        return f"Fout bij ophalen Exact data: {str(e)}", 500


@app.route("/upload-pdfs", methods=["POST"])
def upload_pdfs():
    uploaded_files = request.files.getlist("files")
    pdf_results: List[PdfInvoiceResult] = []

    for file in uploaded_files:
        if not file.filename.lower().endswith(".pdf"):
            continue
        pdf_results.append(parse_pdf_invoice(file.filename, file.read()))

    pdf_df = pd.DataFrame(
        [
            {
                "Factuurnummer": r.factuurnummer,
                "AWB nummer": r.awb_nummer,
                "Totaal kg": r.totaal_kg,
                "Charges (EUR)": r.charges_eur,
                "Prijs per kg (EUR)": r.prijs_per_kg_eur,
                "Bestandsnaam": r.bestandsnaam,
                "Status": r.status,
            }
            for r in pdf_results
        ]
    )
    session["pdf_json"] = pdf_df.to_json(orient="records")

    exact_df = session_get_df("exact_json")
    exact_rows = exact_df.to_dict(orient="records") if not exact_df.empty else []

    merged_df = merge_exact_and_pdf(exact_rows, pdf_results)
    session["merged_json"] = merged_df.to_json(orient="records")

    return redirect(url_for("index"))


@app.route("/download-excel")
def download_excel():
    merged_df = session_get_df("merged_json")
    pdf_df = session_get_df("pdf_json")
    exact_df = session_get_df("exact_json")

    if merged_df.empty:
        return "Nog geen data om te downloaden.", 400

    excel_bytes = build_excel_bytes(merged_df, pdf_df, exact_df)

    return send_file(
        io.BytesIO(excel_bytes),
        as_attachment=True,
        download_name="mission_freight_exact_pdf.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/health")
def health():
    return Response("ok", mimetype="text/plain")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "10000"))
    app.run(host="0.0.0.0", port=port, debug=False)
